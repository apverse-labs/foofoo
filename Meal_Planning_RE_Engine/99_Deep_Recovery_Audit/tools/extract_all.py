#!/usr/bin/env python3
"""
Deep Recovery Audit — full source extractor (Phase 1.2).

Extracts EVERYTHING available from canonical .docx / .xlsx / .md:
  DOCX: paragraphs, tables, comments, footnotes/endnotes, headers/footers, core props.
  XLSX: sheet metadata (incl hidden), rows, cols, formulas, cell comments/notes,
        data validations, named ranges, merged cells, hidden rows/cols, hyperlinks.
  MD:   raw copy + line/heading counts.

Dependency-free for DOCX (zipfile + ElementTree); openpyxl for XLSX.
Writes JSON artifacts + a run log. No network, no DB, read-only on sources.
"""
import os, sys, json, zipfile, hashlib
import xml.etree.ElementTree as ET
import openpyxl

REPO = "/home/user/foofoo"
DOCS = os.path.join(REPO, "Meal_Planning_RE_Engine/Meal_Planning_RE_Technical_Docs_v1")
OUT  = os.path.join(REPO, "Meal_Planning_RE_Engine/99_Deep_Recovery_Audit/01_SOURCE_EXTRACTION")
D_DOCX = os.path.join(OUT, "extracted_docx")
D_XLSX = os.path.join(OUT, "extracted_xlsx")
D_MD   = os.path.join(OUT, "extracted_md")
D_WB   = os.path.join(OUT, "extracted_source_workbook")
LOGDIR = os.path.join(OUT, "extraction_logs")

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

def sha(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for b in iter(lambda: f.read(65536), b""):
            h.update(b)
    return h.hexdigest()[:16]

# ---------- DOCX ----------
def _para_text(p):
    return "".join(t.text or "" for t in p.iter(W + "t"))

def extract_docx(path):
    rec = {"file": path, "paragraphs": [], "tables": [], "comments": [],
           "footnotes": [], "headers_footers": [], "core_props": {}, "limitations": []}
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        # main document: paragraphs (in order) + tables
        root = ET.fromstring(z.read("word/document.xml"))
        body = root.find(W + "body")
        for el in body:
            if el.tag == W + "p":
                txt = _para_text(el).strip()
                if txt:
                    rec["paragraphs"].append(txt)
            elif el.tag == W + "tbl":
                tbl = []
                for tr in el.findall(W + "tr"):
                    row = []
                    for tc in tr.findall(W + "tc"):
                        cell = " ".join(_para_text(p).strip() for p in tc.findall(W + "p"))
                        row.append(cell.strip())
                    tbl.append(row)
                rec["tables"].append(tbl)
        # comments
        if "word/comments.xml" in names:
            cr = ET.fromstring(z.read("word/comments.xml"))
            for c in cr.findall(W + "comment"):
                rec["comments"].append({
                    "author": c.get(W + "author"),
                    "date": c.get(W + "date"),
                    "text": "".join(t.text or "" for t in c.iter(W + "t")).strip(),
                })
        # footnotes / endnotes
        for fn in ("word/footnotes.xml", "word/endnotes.xml"):
            if fn in names:
                fr = ET.fromstring(z.read(fn))
                for note in fr.iter(W + ("footnote" if "foot" in fn else "endnote")):
                    txt = "".join(t.text or "" for t in note.iter(W + "t")).strip()
                    if txt:
                        rec["footnotes"].append(txt)
        # headers / footers
        for n in names:
            if n.startswith("word/header") or n.startswith("word/footer"):
                hr = ET.fromstring(z.read(n))
                txt = "".join(t.text or "" for t in hr.iter(W + "t")).strip()
                if txt:
                    rec["headers_footers"].append({"part": n, "text": txt})
        # core properties
        if "docProps/core.xml" in names:
            cp = ET.fromstring(z.read("docProps/core.xml"))
            for el in cp.iter():
                tag = el.tag.split("}")[-1]
                if el.text and el.text.strip():
                    rec["core_props"][tag] = el.text.strip()
    return rec

# ---------- XLSX ----------
def extract_xlsx(path):
    rec = {"file": path, "named_ranges": {}, "sheets": [], "limitations": []}
    # values pass (data_only=True) and formula pass (data_only=False)
    wb_v = openpyxl.load_workbook(path, read_only=False, data_only=True)
    wb_f = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        for name, dest in (wb_f.defined_names.items() if hasattr(wb_f.defined_names, "items")
                           else [(d.name, d) for d in wb_f.defined_names.definedName]):
            try:
                rec["named_ranges"][name] = str(getattr(dest, "value", dest))
            except Exception:
                rec["named_ranges"][name] = "?"
    except Exception as e:
        rec["limitations"].append(f"named_ranges: {e}")

    for ws_f in wb_f.worksheets:
        ws_v = wb_v[ws_f.title]
        s = {"name": ws_f.title, "state": ws_f.sheet_state,
             "max_row": ws_f.max_row, "max_col": ws_f.max_column,
             "headers": [], "rows": [], "formulas": [], "comments": [],
             "data_validations": [], "merged_cells": [], "hidden_rows": [],
             "hidden_cols": [], "hyperlinks": []}
        # headers (first row)
        if ws_v.max_row >= 1:
            s["headers"] = [c.value for c in ws_v[1]]
        # rows (values)
        for r in ws_v.iter_rows(values_only=True):
            if any(v is not None for v in r):
                s["rows"].append(list(r))
        # formulas + comments + hyperlinks (formula pass)
        for row in ws_f.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    s["formulas"].append({"cell": c.coordinate, "formula": c.value})
                if c.comment is not None:
                    s["comments"].append({"cell": c.coordinate,
                                          "author": c.comment.author,
                                          "text": (c.comment.text or "").strip()})
                if c.hyperlink is not None:
                    s["hyperlinks"].append({"cell": c.coordinate, "target": c.hyperlink.target})
        # data validations
        try:
            for dv in ws_f.data_validations.dataValidation:
                s["data_validations"].append({"type": dv.type, "formula1": dv.formula1,
                                              "ranges": str(dv.sqref)})
        except Exception as e:
            s["limitations"].append(f"dv: {e}")
        # merged cells
        s["merged_cells"] = [str(m) for m in ws_f.merged_cells.ranges]
        # hidden rows / cols
        for idx, dim in ws_f.row_dimensions.items():
            if dim.hidden:
                s["hidden_rows"].append(idx)
        for col, dim in ws_f.column_dimensions.items():
            if dim.hidden:
                s["hidden_cols"].append(col)
        rec["sheets"].append(s)
    wb_v.close(); wb_f.close()
    return rec

# ---------- driver ----------
def main():
    for d in (D_DOCX, D_XLSX, D_MD, D_WB, LOGDIR):
        os.makedirs(d, exist_ok=True)
    log = []
    workbook_path = os.path.join(DOCS, "09_Source_Data/Indian_Meal_Cohort_Persona_DB_v3.xlsx")

    for dirpath, _, files in os.walk(DOCS):
        for fn in sorted(files):
            full = os.path.join(dirpath, fn)
            ext = fn.lower().rsplit(".", 1)[-1]
            base = fn.rsplit(".", 1)[0]
            entry = {"file": os.path.relpath(full, REPO), "ext": ext, "sha": sha(full),
                     "status": "ok", "limitation": ""}
            try:
                if ext == "docx":
                    rec = extract_docx(full)
                    json.dump(rec, open(os.path.join(D_DOCX, base + ".json"), "w"), indent=1, default=str)
                    entry.update(paragraphs=len(rec["paragraphs"]), tables=len(rec["tables"]),
                                 comments=len(rec["comments"]), footnotes=len(rec["footnotes"]),
                                 headers_footers=len(rec["headers_footers"]))
                elif ext == "xlsx":
                    rec = extract_xlsx(full)
                    dest = D_WB if full == workbook_path else D_XLSX
                    json.dump(rec, open(os.path.join(dest, base + ".json"), "w"), indent=1, default=str)
                    entry.update(sheets=len(rec["sheets"]),
                                 rows=sum(len(s["rows"]) for s in rec["sheets"]),
                                 formulas=sum(len(s["formulas"]) for s in rec["sheets"]),
                                 comments=sum(len(s["comments"]) for s in rec["sheets"]),
                                 hidden_sheets=sum(1 for s in rec["sheets"] if s["state"] != "visible"),
                                 data_validations=sum(len(s["data_validations"]) for s in rec["sheets"]),
                                 named_ranges=len(rec["named_ranges"]),
                                 merged=sum(len(s["merged_cells"]) for s in rec["sheets"]))
                elif ext == "md":
                    txt = open(full, encoding="utf-8").read()
                    open(os.path.join(D_MD, fn), "w", encoding="utf-8").write(txt)
                    entry.update(lines=txt.count("\n") + 1,
                                 headings=sum(1 for l in txt.splitlines() if l.lstrip().startswith("#")))
                else:
                    entry["status"] = "skipped"; entry["limitation"] = "non-canonical ext"
            except Exception as e:
                entry["status"] = "FAILED"; entry["limitation"] = repr(e)
            log.append(entry)

    json.dump(log, open(os.path.join(LOGDIR, "extraction_run_log.json"), "w"), indent=1, default=str)
    ok = sum(1 for e in log if e["status"] == "ok")
    failed = [e for e in log if e["status"] == "FAILED"]
    print(f"Extracted OK: {ok} / {len(log)} files")
    for e in log:
        extra = {k: v for k, v in e.items() if k not in ("file", "ext", "sha", "status", "limitation")}
        print(f"  [{e['status']}] {os.path.basename(e['file'])}: {extra}")
    if failed:
        print("\nFAILURES:")
        for e in failed:
            print(f"  {e['file']}: {e['limitation']}")
        sys.exit(1)

if __name__ == "__main__":
    main()
